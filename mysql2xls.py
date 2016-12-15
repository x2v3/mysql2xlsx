from openpyxl import Workbook
import pymysql

class Mysql2xls:
    database=None
    workbook=None

    def datatable2worksheetByName(self,datatableName):
        tables=database.get_tables()
        t=[tb for tb in tables if tb.name==tablename][0]


    def db2workbook(self):
        for t in self.database.get_tables():
            self.datatable2worksheet(t)

    def datatable2worksheet(self,datatable):
        if self.workbook is None: 
            self.workbook=Workbook()
        ws=self.workbook.create_sheet()
        t=datatable
        ws.title=t.name

        #fill worksheet
        columnIndex=1
        rowIndex=1
        #generate header
        for c in t.get_columns():
            ws.cell(row=rowIndex,column=columnIndex,value=c)
            columnIndex+=1
        rowIndex += 1

        #generate data rows
        for r in t.rows:
            columnIndex=1
            for d in r:
                ws.cell(row=rowIndex,column=columnIndex,value=d)
                columnIndex+=1
            rowIndex+=1

    def saveworkbook(self,path):
        self.workbook.save(path)



